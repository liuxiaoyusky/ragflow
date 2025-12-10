#!/usr/bin/env python3
"""
Factsheet PDF 处理器

使用 MinerU API 解析 PDF，生成独立的section markdown文件：
- 每个section一个markdown文件
- 支持批量处理27个PDF
- 支持自动上传到RagFlow并添加keywords

用法：
    python factsheet_processor.py input.pdf                    # 处理单个文件
    python factsheet_processor.py input.pdf -o output_dir      # 指定输出目录
    python factsheet_processor.py --batch input_dir            # 批量处理目录
    python factsheet_processor.py --upload output_dir          # 上传到RagFlow
"""

import argparse
import json
import os
import re
import tempfile
import zipfile
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# MinerU API 配置
MINERU_API = os.environ.get("MINERU_API", "http://10.1.9.133:9987")


# ============ MinerU API 调用 ============

def parse_pdf_with_mineru(pdf_path: str, output_dir: str = None) -> dict:
    """
    使用 MinerU API 解析 PDF
    返回 content_list.json 的内容
    """
    pdf_path = Path(pdf_path)
    pdf_name = pdf_path.stem.replace(" ", "")
    
    if output_dir is None:
        output_dir = tempfile.mkdtemp(prefix="mineru_")
    
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    # 调用 MinerU API
    print(f"[MinerU] 解析 {pdf_path.name}...")
    
    files = {
        "files": (pdf_name + ".pdf", open(pdf_path, "rb"), "application/pdf")
    }
    
    data = {
        "output_dir": "./output",
        "backend": "vlm-vllm-async-engine",  # 使用VLM backend提高表格识别精度
        "parse_method": "ocr",                # 强制OCR模式，更精确
        "lang_list": '["en"]',               # 指定英语
        "formula_enable": "true",
        "table_enable": "true",
        "return_md": "true",
        "return_middle_json": "true",
        "return_content_list": "true",
        "return_images": "true",
        "response_format_zip": "true",
    }
    
    try:
        response = requests.post(
            f"{MINERU_API}/file_parse",
            files=files,
            data=data,
            timeout=1800
        )
        response.raise_for_status()
        
        if response.headers.get("Content-Type") == "application/zip":
            # 解压 zip 文件
            zip_path = output_path / "output.zip"
            with open(zip_path, "wb") as f:
                f.write(response.content)
            
            with zipfile.ZipFile(zip_path, "r") as zip_ref:
                zip_ref.extractall(output_path)
            
            # 查找 content_list.json
            content_list_path = None
            for root, dirs, files in os.walk(output_path):
                for file in files:
                    if file.endswith("_content_list.json"):
                        content_list_path = Path(root) / file
                        break
            
            if content_list_path and content_list_path.exists():
                with open(content_list_path, "r", encoding="utf-8") as f:
                    content_list = json.load(f)
                print(f"[MinerU] 解析完成，共 {len(content_list)} 个blocks")
                return {
                    "content_list": content_list,
                    "output_dir": str(output_path),
                    "pdf_name": pdf_name
                }
            else:
                raise FileNotFoundError("找不到 content_list.json")
        else:
            raise RuntimeError(f"MinerU 返回非 zip 格式: {response.headers.get('Content-Type')}")
            
    except Exception as e:
        print(f"[MinerU] 解析失败: {e}")
        raise


# ============ 内容分类和处理 ============

# 章节标题映射
SECTION_TITLES = {
    "investment objective": "Investment objective",
    "performance since launch": "Performance since launch",
    "performance update": "Performance update",
    "navs & codes": "NAVs & codes",
    "dividend information": "Dividend information",
    "top holdings - equities": "Top holdings - equities",
    "top holdings – equities": "Top holdings - equities",
    "top holdings - fixed income": "Top holdings - fixed income",
    "top holdings – fixed income": "Top holdings - fixed income",
    "monthly performance": "Monthly performance",
    "portfolio characteristics": "Portfolio characteristics",
    "asset type by geography": "Asset type by geography",
    "asset type by sector": "Asset type by sector",
    "credit ratings": "Credit ratings",
    "fund facts": "Fund facts",
    "fee structure": "Fee structure",
    "subscription information": "Subscription information",
    "senior investment staff": "Senior investment staff",
    "key fund and corporate awards": "Key fund and corporate awards",
}


def detect_table_title(table_content: str, prev_text: str = "") -> str:
    """
    检测表格的标题
    优先从表格后面的caption检测，其次从前面的文本检测
    """
    content_lower = table_content.lower()
    
    # 检测特定的表格类型
    if "taiwan semiconductor" in content_lower or "sk hynix" in content_lower:
        if "name" in content_lower and "industry" in content_lower:
            return "Top holdings - equities"
    
    if "fortune star" in content_lower or "sumitomo mitsui" in content_lower:
        if "sector" in content_lower:
            return "Top holdings - fixed income"
    
    if "dividend amount" in content_lower or "annualized yield" in content_lower:
        return "Dividend information"
    
    if "year-to-date" in content_lower or "one month" in content_lower:
        return "Performance update"
    
    if "nav" in content_lower and "isin" in content_lower:
        return "NAVs & codes"
    
    if "portfolio yield" in content_lower or "annualized volatility" in content_lower:
        return "Portfolio characteristics"
    
    if "jan" in content_lower and "feb" in content_lower and "annual" in content_lower:
        return "Monthly performance"
    
    if "equities" in content_lower and "fixed income" in content_lower and "total" in content_lower:
        return "Asset type by sector"
    
    # 从前面的文本检测
    if prev_text:
        prev_lower = prev_text.lower()
        for key, title in SECTION_TITLES.items():
            if key in prev_lower:
                return title
    
    return "Table"


def html_table_to_dataframe(html_table: str) -> pd.DataFrame:
    """
    将 HTML 表格转换为 DataFrame
    """
    try:
        # 使用 pandas 解析 HTML 表格
        dfs = pd.read_html(html_table)
        if dfs:
            return dfs[0]
    except Exception as e:
        print(f"[警告] 无法解析 HTML 表格: {e}")
    
    return None


def detect_section_title(text: str) -> str:
    """
    检测文本属于哪个章节
    """
    text_lower = text.lower()
    for key, title in SECTION_TITLES.items():
        if key in text_lower:
            return title
    return None


# ============ 输出生成 ============

def process_content_list(content_list: list, pdf_name: str) -> dict:
    """
    处理 content_list，分离表格和文本
    
    返回:
    {
        "tables": [(title, dataframe), ...],
        "texts": [(section_title, text), ...],
        "metadata": {...}
    }
    """
    tables = []
    texts = []
    
    # 提取元数据
    metadata = extract_metadata(pdf_name)
    
    prev_text = ""
    current_section = None  # 跟踪当前section，用于继承
    
    for i, item in enumerate(content_list):
        item_type = item.get("type", "")
        
        if item_type == "table":
            # 处理表格
            table_body = item.get("table_body", "")
            table_caption = " ".join(item.get("table_caption", []))
            table_footnote = " ".join(item.get("table_footnote", []))
            
            full_content = table_body + "\n" + table_caption + "\n" + table_footnote
            
            # 检测表格标题
            title = detect_table_title(full_content, prev_text)
            
            # 如果有 caption，优先使用
            if table_caption:
                for key, section_title in SECTION_TITLES.items():
                    if key in table_caption.lower():
                        title = section_title
                        break
            
            # 如果没有检测到标题，使用当前section
            if title == "Table" and current_section:
                title = current_section
            
            # 转换为 DataFrame
            if table_body and "<table>" in table_body.lower():
                df = html_table_to_dataframe(table_body)
                if df is not None:
                    tables.append((title, df, table_caption))
            
        elif item_type == "text":
            # 处理文本
            text = item.get("text", "")
            if text.strip():
                detected_section = detect_section_title(text)
                
                # 如果检测到新的section标题，更新current_section
                if detected_section:
                    current_section = detected_section
                
                # 使用检测到的section或继承当前section
                section_title = detected_section if detected_section else current_section
                texts.append((section_title, text))
                prev_text = text
        
        elif item_type == "image":
            # 图片暂时跳过，或者可以添加占位符
            captions = item.get("image_caption", [])
            if captions:
                texts.append(("Image", " ".join(captions)))
    
    return {
        "tables": tables,
        "texts": texts,
        "metadata": metadata
    }


def extract_metadata(pdf_name: str) -> dict:
    """
    从文件名提取元数据
    例如: VP_Asian Income Fund-202506-Eng.pdf
    """
    metadata = {
        "fund_type": "",
        "report_month": "",
        "language": "English"
    }
    
    # 提取基金类型 (支持有空格和无空格两种格式)
    if "Asian Income" in pdf_name or "AsianIncome" in pdf_name:
        metadata["fund_type"] = "Asian Income Fund"
    elif "High-Dividend" in pdf_name or "High Dividend" in pdf_name or "HighDividend" in pdf_name:
        metadata["fund_type"] = "High Dividend Fund"
    elif "Classic" in pdf_name:
        metadata["fund_type"] = "Classic Fund"
    
    # 提取报告月份
    match = re.search(r"(\d{6})", pdf_name)
    if match:
        yyyymm = match.group(1)
        year = yyyymm[:4]
        month = yyyymm[4:6]
        month_name = datetime(int(year), int(month), 1).strftime("%B %Y")
        metadata["report_month"] = month_name
    
    return metadata


def generate_table_markdown(tables: list, output_dir: str, metadata: dict, base_name: str):
    """
    为每个表格生成单独的 Markdown 文件
    ragflow对markdown有优化，会按section分块
    整个文件会成为一个chunk，标题自动可搜索
    """
    generated_files = []
    title_count = {}
    
    fund_type = metadata.get('fund_type', '')
    report_month = metadata.get('report_month', '')
    
    for title, df, caption in tables:
        # 生成唯一的文件名
        safe_title = re.sub(r'[\\/*?:\[\]<>|"\s]+', '_', title)[:30]
        if safe_title in title_count:
            title_count[safe_title] += 1
            file_suffix = f"_{title_count[safe_title]}"
        else:
            title_count[safe_title] = 1
            file_suffix = ""
        
        file_name = f"{base_name}_{safe_title}{file_suffix}.md"
        file_path = os.path.join(output_dir, file_name)
        
        # 生成Markdown内容
        md_content = []
        
        # 标题（作为关键词）
        md_content.append(f"# {title}")
        md_content.append("")
        
        # 元数据
        if fund_type:
            md_content.append(f"**Fund:** {fund_type}")
        if report_month:
            md_content.append(f"**Month:** {report_month}")
        if caption:
            md_content.append(f"**Note:** {caption}")
        md_content.append("")
        
        # 转换DataFrame为Markdown表格
        md_table = df.to_markdown(index=False)
        md_content.append(md_table)
        
        # 写入文件
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(md_content))
        
        generated_files.append(file_path)
    
    print(f"[Markdown] 生成 {len(generated_files)} 个独立表格文件到 {output_dir}")
    return generated_files


def generate_excel(tables: list, output_dir: str, metadata: dict, base_name: str):
    """
    为每个表格生成单独的 Excel 文件
    这样每个表格在上传后会成为独立的chunk
    """
    generated_files = []
    title_count = {}
    
    for title, df, caption in tables:
        # 生成唯一的文件名
        safe_title = re.sub(r'[\\/*?:\[\]<>|"\s]+', '_', title)[:30]
        if safe_title in title_count:
            title_count[safe_title] += 1
            file_suffix = f"_{title_count[safe_title]}"
        else:
            title_count[safe_title] = 1
            file_suffix = ""
        
        file_name = f"{base_name}_{safe_title}{file_suffix}.xlsx"
        file_path = os.path.join(output_dir, file_name)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # 添加元数据作为注释（让ragflow可以搜索到）
        ws.cell(row=1, column=1, value=f"Fund: {metadata.get('fund_type', '')}")
        ws.cell(row=2, column=1, value=f"Month: {metadata.get('report_month', '')}")
        ws.cell(row=3, column=1, value=f"Section: {title}")
        if caption:
            ws.cell(row=4, column=1, value=f"Caption: {caption}")
        
        # 写入数据（从第6行开始）
        start_row = 6
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        wb.save(file_path)
        generated_files.append(file_path)
    
    print(f"[Excel] 生成 {len(generated_files)} 个独立表格文件到 {output_dir}")
    return generated_files


def generate_markdown(texts: list, output_path: str, metadata: dict):
    """
    生成 Markdown 文件
    按章节组织文本
    """
    md_content = []
    
    # 添加头部元数据
    md_content.append(f"# {metadata.get('fund_type', 'Fund')} - {metadata.get('report_month', '')}")
    md_content.append("")
    md_content.append(f"**Fund Type:** {metadata.get('fund_type', '')}")
    md_content.append(f"**Report Month:** {metadata.get('report_month', '')}")
    md_content.append("")
    md_content.append("---")
    md_content.append("")
    
    current_section = None
    
    for section_title, text in texts:
        # 如果是新章节，添加标题
        if section_title and section_title != current_section:
            md_content.append(f"## {section_title}")
            md_content.append("")
            current_section = section_title
        
        # 添加文本内容
        md_content.append(text)
        md_content.append("")
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(md_content))
    
    print(f"[Markdown] 生成 {output_path}，共 {len(texts)} 段文本")


def generate_complete_markdown(tables: list, texts: list, metadata: dict) -> str:
    """
    生成完整的Markdown文件内容（包含所有表格和文本）
    用---分隔每个section
    """
    fund_type = metadata.get('fund_type', '')
    report_month = metadata.get('report_month', '')
    
    md_content = []
    
    # 头部信息
    md_content.append(f"# {fund_type} - {report_month}")
    md_content.append("")
    md_content.append(f"**Fund:** {fund_type}  ")
    md_content.append(f"**Report Month:** {report_month}  ")
    md_content.append("")
    md_content.append("---")
    md_content.append("")
    
    # 按标准顺序组织sections
    section_order = [
        "Investment objective",
        "Performance update",
        "NAVs & codes",
        "Dividend information",
        "Top holdings - equities",
        "Top holdings - fixed income",
        "Monthly performance",
        "Portfolio characteristics",
        "Asset type by geography",
        "Asset type by sector",
        "Credit ratings",
        "Fee structure",
        "Fund facts",
        "Risk Disclosure",
    ]
    
    # 将表格按标题组织
    table_by_title = {}
    for title, df, caption in tables:
        if title not in table_by_title:
            table_by_title[title] = []
        table_by_title[title].append((df, caption))
    
    # 将文本按标题组织
    text_by_section = {}
    for section_title, text in texts:
        if section_title:
            if section_title not in text_by_section:
                text_by_section[section_title] = []
            text_by_section[section_title].append(text)
    
    # 按顺序生成sections
    generated_sections = set()
    
    for section_title in section_order:
        has_content = False
        section_content = []
        
        # 添加标题
        section_content.append(f"## {section_title}")
        section_content.append("")
        
        # 添加文本
        if section_title in text_by_section:
            for text in text_by_section[section_title]:
                section_content.append(text)
                section_content.append("")
            has_content = True
        
        # 添加表格
        if section_title in table_by_title:
            for df, caption in table_by_title[section_title]:
                if caption:
                    section_content.append(f"*{caption}*")
                    section_content.append("")
                md_table = df.to_markdown(index=False)
                section_content.append(md_table)
                section_content.append("")
            has_content = True
        
        if has_content:
            md_content.extend(section_content)
            md_content.append("---")
            md_content.append("")
            generated_sections.add(section_title)
    
    # 添加未分类的表格
    for title, items in table_by_title.items():
        if title not in generated_sections and title != "Table":
            md_content.append(f"## {title}")
            md_content.append("")
            for df, caption in items:
                if caption:
                    md_content.append(f"*{caption}*")
                    md_content.append("")
                md_table = df.to_markdown(index=False)
                md_content.append(md_table)
                md_content.append("")
            md_content.append("---")
            md_content.append("")
    
    return "\n".join(md_content)


def split_to_sections(complete_md: str, metadata: dict, output_dir: str, base_name: str) -> list:
    """
    将完整的markdown按---分割成独立的section文件
    返回: [(file_path, title, keywords), ...]
    """
    fund_type = metadata.get('fund_type', '')
    report_month = metadata.get('report_month', '')
    
    # 基础元数据（添加到每个section）
    base_meta = f"**Fund:** {fund_type}  \n**Report Month:** {report_month}\n\n"
    
    # 用---分割
    sections = complete_md.split('\n---\n')
    
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    generated_files = []
    
    for i, section in enumerate(sections):
        section = section.strip()
        if not section:
            continue
        
        # 提取标题
        title_match = re.search(r'^#+ (.+)$', section, re.MULTILINE)
        if title_match:
            title = title_match.group(1).strip()
        else:
            title = f"Section_{i+1}"
        
        # 跳过头部信息section（只有基金名和月份）
        if i == 0 and len(section) < 300 and "---" not in section:
            continue
        
        # 生成安全的文件名
        safe_title = re.sub(r'[^\w\s-]', '', title)
        safe_title = re.sub(r'\s+', '_', safe_title)[:50]
        
        filename = f"{base_name}_{safe_title}.md"
        filepath = output_path / filename
        
        # 如果section不包含Fund信息，添加基础元数据
        if fund_type and fund_type not in section:
            section = base_meta + section
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(section)
        
        # 生成keywords
        keywords = generate_keywords(title, metadata)
        
        generated_files.append({
            "path": str(filepath),
            "filename": filename,
            "title": title,
            "keywords": keywords
        })
    
    return generated_files


def generate_keywords(title: str, metadata: dict) -> list:
    """
    从标题和元数据生成keywords
    """
    keywords = []
    
    # 添加标题作为keyword
    keywords.append(title)
    
    # 特殊处理一些标题
    title_lower = title.lower()
    if "top holdings" in title_lower:
        keywords.append("Top holdings")
        if "equities" in title_lower:
            keywords.append("equities")
        if "fixed income" in title_lower:
            keywords.append("fixed income")
    elif "dividend" in title_lower:
        keywords.append("dividend")
    elif "performance" in title_lower:
        keywords.append("performance")
    elif "nav" in title_lower:
        keywords.append("NAV")
    elif "asset" in title_lower:
        if "geography" in title_lower:
            keywords.append("geography")
        if "sector" in title_lower:
            keywords.append("sector")
    elif "credit" in title_lower:
        keywords.append("credit rating")
    elif "fee" in title_lower:
        keywords.append("fee")
    
    # 添加基金类型和月份
    if metadata.get('fund_type'):
        keywords.append(metadata['fund_type'])
    if metadata.get('report_month'):
        keywords.append(metadata['report_month'])
    
    return list(set(keywords))


# ============ 主程序 ============

def process_factsheet(pdf_path: str, output_dir: str = None) -> list:
    """
    处理单个 factsheet PDF
    返回: 生成的section文件列表 [{path, filename, title, keywords}, ...]
    """
    pdf_path = Path(pdf_path)
    
    if output_dir is None:
        output_dir = pdf_path.parent / "sections"
    
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"\n{'='*60}")
    print(f"处理: {pdf_path.name}")
    print(f"{'='*60}")
    
    # 1. 调用 MinerU 解析
    result = parse_pdf_with_mineru(str(pdf_path))
    content_list = result["content_list"]
    pdf_name = result["pdf_name"]
    
    # 2. 处理内容
    processed = process_content_list(content_list, pdf_name)
    metadata = processed["metadata"]
    
    # 3. 生成完整的markdown
    complete_md = generate_complete_markdown(
        processed["tables"],
        processed["texts"],
        metadata
    )
    
    # 4. 按section拆分成独立文件
    # 生成基础文件名：VP_Asian_Income_Fund_202508
    fund_short = metadata.get('fund_type', 'Fund').replace(' ', '_')
    month_match = re.search(r'(\w+)\s+(\d{4})', metadata.get('report_month', ''))
    if month_match:
        month_short = f"{month_match.group(2)}{month_match.group(1)[:3]}"
    else:
        month_short = re.sub(r'\s+', '', metadata.get('report_month', ''))
    base_name = f"VP_{fund_short}_{month_short}"
    
    section_files = split_to_sections(
        complete_md,
        metadata,
        str(output_dir),
        base_name
    )
    
    print(f"\n输出目录: {output_dir}")
    print(f"生成 {len(section_files)} 个section文件:")
    for f in section_files:
        print(f"  - {f['filename']} [{', '.join(f['keywords'][:3])}...]")
    
    return section_files


def upload_to_ragflow(section_files: list, dataset_id: str, api_key: str, base_url: str = "https://10.1.9.133:8443"):
    """
    上传section文件到RagFlow并添加keywords
    """
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    headers = {"Authorization": f"Bearer {api_key}"}
    uploaded = []
    
    print(f"\n上传 {len(section_files)} 个文件到 RagFlow...")
    
    for f in section_files:
        filepath = f["path"]
        keywords = f["keywords"]
        
        # 1. 上传文件
        with open(filepath, "rb") as file:
            files = {"file": (Path(filepath).name, file, "text/markdown")}
            response = requests.post(
                f"{base_url}/api/v1/datasets/{dataset_id}/documents",
                headers=headers,
                files=files,
                verify=False
            )
        
        if response.status_code != 200:
            print(f"  [错误] 上传 {f['filename']} 失败: {response.text}")
            continue
        
        result = response.json()
        if result.get("code") != 0:
            print(f"  [错误] 上传 {f['filename']} 失败: {result.get('message')}")
            continue
        
        doc_id = result.get("data", [{}])[0].get("id")
        if doc_id:
            uploaded.append({
                "doc_id": doc_id,
                "filename": f["filename"],
                "keywords": keywords
            })
            print(f"  ✓ {f['filename']} -> {doc_id}")
    
    if not uploaded:
        print("没有文件成功上传")
        return []
    
    # 2. 触发解析
    doc_ids = [u["doc_id"] for u in uploaded]
    print(f"\n触发解析 {len(doc_ids)} 个文档...")
    
    response = requests.post(
        f"{base_url}/api/v1/datasets/{dataset_id}/chunks",
        headers={**headers, "Content-Type": "application/json"},
        json={"document_ids": doc_ids},
        verify=False
    )
    
    if response.status_code == 200 and response.json().get("code") == 0:
        print("  ✓ 解析任务已提交")
    else:
        print(f"  [错误] 触发解析失败: {response.text}")
    
    # 3. 等待解析完成并添加keywords
    print("\n等待解析完成...")
    import time
    time.sleep(10)  # 等待解析
    
    # 为每个chunk添加keywords
    print("\n添加keywords...")
    for u in uploaded:
        doc_id = u["doc_id"]
        keywords = u["keywords"]
        
        # 获取chunks
        response = requests.get(
            f"{base_url}/api/v1/datasets/{dataset_id}/documents/{doc_id}/chunks?page=1&page_size=10",
            headers=headers,
            verify=False
        )
        
        if response.status_code != 200:
            continue
        
        chunks = response.json().get("data", {}).get("chunks", [])
        for chunk in chunks:
            chunk_id = chunk.get("id")
            if chunk_id:
                # 更新keywords
                response = requests.put(
                    f"{base_url}/api/v1/datasets/{dataset_id}/documents/{doc_id}/chunks/{chunk_id}",
                    headers={**headers, "Content-Type": "application/json"},
                    json={"important_keywords": keywords},
                    verify=False
                )
                if response.status_code == 200:
                    print(f"  ✓ {u['filename']}: {keywords[:3]}...")
    
    return uploaded


def main():
    global MINERU_API
    
    parser = argparse.ArgumentParser(description="Factsheet PDF 处理器")
    parser.add_argument("input", nargs="?", help="输入 PDF 文件路径")
    parser.add_argument("-o", "--output", help="输出目录")
    parser.add_argument("--batch", help="批量处理目录中的所有PDF")
    parser.add_argument("--api", default=MINERU_API, help=f"MinerU API 地址 (默认: {MINERU_API})")
    parser.add_argument("--upload", help="上传指定目录的markdown文件到RagFlow")
    parser.add_argument("--dataset-id", help="RagFlow dataset ID")
    parser.add_argument("--api-key", default="ragflow-NB-Mo_q3z2SaAgPl_Bn9t6lgysmj6cj-WPWEdUV-7Iw", help="RagFlow API key")
    parser.add_argument("--ragflow-url", default="https://10.1.9.133:8443", help="RagFlow URL")
    
    args = parser.parse_args()
    MINERU_API = args.api
    
    all_section_files = []
    
    if args.batch:
        # 批量处理
        input_dir = Path(args.batch)
        output_base = Path(args.output) if args.output else input_dir.parent / "factsheets_processed" / "sections"
        
        pdf_files = sorted(input_dir.glob("*.pdf"))
        print(f"找到 {len(pdf_files)} 个 PDF 文件")
        
        for i, pdf_file in enumerate(pdf_files, 1):
            print(f"\n[{i}/{len(pdf_files)}]", end="")
            try:
                section_files = process_factsheet(str(pdf_file), str(output_base))
                all_section_files.extend(section_files)
            except Exception as e:
                print(f"[错误] 处理 {pdf_file.name} 失败: {e}")
                import traceback
                traceback.print_exc()
        
        print(f"\n{'='*60}")
        print(f"批量处理完成！共生成 {len(all_section_files)} 个section文件")
        print(f"输出目录: {output_base}")
        
        # 保存文件列表（用于后续上传）
        manifest_path = output_base / "manifest.json"
        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(all_section_files, f, ensure_ascii=False, indent=2)
        print(f"文件清单: {manifest_path}")
    
    elif args.upload:
        # 上传到RagFlow
        upload_dir = Path(args.upload)
        manifest_path = upload_dir / "manifest.json"
        
        if manifest_path.exists():
            with open(manifest_path, "r", encoding="utf-8") as f:
                section_files = json.load(f)
        else:
            # 扫描目录中的md文件
            section_files = []
            for md_file in upload_dir.glob("*.md"):
                # 从文件名提取信息
                title = md_file.stem.split("_")[-1].replace("_", " ")
                section_files.append({
                    "path": str(md_file),
                    "filename": md_file.name,
                    "title": title,
                    "keywords": [title]
                })
        
        if not args.dataset_id:
            print("错误: 请指定 --dataset-id")
            return
        
        upload_to_ragflow(
            section_files,
            args.dataset_id,
            args.api_key,
            args.ragflow_url
        )
    
    elif args.input:
        # 处理单个文件
        output_dir = args.output if args.output else str(Path(args.input).parent / "sections")
        section_files = process_factsheet(args.input, output_dir)
        
        # 保存文件列表
        output_path = Path(output_dir)
        manifest_path = output_path / "manifest.json"
        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(section_files, f, ensure_ascii=False, indent=2)
    
    else:
        parser.print_help()


if __name__ == "__main__":
    main()

